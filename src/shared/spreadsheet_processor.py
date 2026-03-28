"""
Spreadsheet Processor - Parse and store spreadsheets for hybrid Postgres+Qdrant access.

Supports: XLSX, XLS, CSV, TSV files
Storage: Postgres for structured queries, Qdrant for semantic search
"""

import os
import csv
import hashlib
import json
import uuid
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass, field
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


@dataclass
class ColumnInfo:
    name: str
    index: int
    detected_type: str = "string"
    sample_values: List[Any] = field(default_factory=list)


@dataclass
class SpreadsheetInfo:
    spreadsheet_id: str
    name: str
    source_path: str
    file_type: str
    sheet_name: Optional[str] = None
    sheet_index: int = 0
    row_count: int = 0
    column_count: int = 0
    column_headers: List[str] = field(default_factory=list)
    column_types: Dict[str, str] = field(default_factory=dict)
    status: str = "pending"
    error_message: Optional[str] = None


@dataclass
class RowData:
    row_index: int
    row_data: Dict[str, Any]
    row_hash: str


class SpreadsheetProcessor:
    """Parse and store spreadsheets for hybrid Postgres+Qdrant access."""
    
    SUPPORTED_EXTENSIONS = {'.xlsx', '.xls', '.csv', '.tsv'}
    
    def __init__(self, max_rows: int = 100000, sample_size: int = 100):
        self.max_rows = max_rows
        self.sample_size = sample_size
        self._check_dependencies()
    
    def _check_dependencies(self):
        if not OPENPYXL_AVAILABLE:
            print("Warning: openpyxl not installed. XLSX support disabled.")
        if not XLRD_AVAILABLE:
            print("Warning: xlrd not installed. XLS (old Excel) support disabled.")
    
    def get_file_type(self, file_path: str) -> Optional[str]:
        ext = os.path.splitext(file_path)[1].lower()
        if ext in self.SUPPORTED_EXTENSIONS:
            return ext.lstrip('.')
        return None
    
    def _detect_column_type(self, values: List[Any]) -> str:
        if not values:
            return "string"
        
        types = {"number": 0, "boolean": 0, "date": 0, "string": 0}
        
        for v in values[:self.sample_size]:
            if v is None or v == "":
                continue
            if isinstance(v, bool):
                types["boolean"] += 1
            elif isinstance(v, (int, float)):
                types["number"] += 1
            elif isinstance(v, datetime):
                types["date"] += 1
            else:
                if isinstance(v, str):
                    v_lower = v.lower().strip()
                    if v_lower in ("true", "false", "yes", "no", "1", "0"):
                        types["boolean"] += 1
                    else:
                        try:
                            float(v.replace(",", ""))
                            types["number"] += 1
                        except (ValueError, AttributeError):
                            types["string"] += 1
                else:
                    types["string"] += 1
        
        if types["number"] > types["string"] and types["number"] > 0:
            return "number"
        if types["boolean"] > types["string"] and types["boolean"] > 0:
            return "boolean"
        if types["date"] > 0:
            return "date"
        return "string"
    
    def _compute_row_hash(self, row_data: Dict[str, Any]) -> str:
        content = json.dumps(row_data, sort_keys=True, default=str)
        return hashlib.sha256(content.encode()).hexdigest()[:16]
    
    def parse_xlsx(self, file_path: str, sheet_index: int = 0) -> Tuple[SpreadsheetInfo, List[RowData]]:
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for XLSX files. Install with: pip install openpyxl")
        
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        
        if sheet_index >= len(sheet_names):
            sheet_index = 0
        
        sheet = workbook[sheet_names[sheet_index]]
        
        rows_data = []
        headers = []
        column_samples: Dict[int, List[Any]] = {}
        row_idx = 0
        
        for row in sheet.iter_rows():
            if row_idx >= self.max_rows:
                break
            
            values = [cell.value for cell in row]
            
            if row_idx == 0:
                if all(v is not None for v in values[:min(5, len(values))]):
                    headers = [str(v).strip() if v else f"column_{i}" for i, v in enumerate(values)]
                else:
                    headers = [f"column_{i}" for i in range(len(values))]
            else:
                row_dict = {}
                for i, v in enumerate(values):
                    col_name = headers[i] if i < len(headers) else f"column_{i}"
                    row_dict[col_name] = v
                    if i not in column_samples:
                        column_samples[i] = []
                    if len(column_samples[i]) < self.sample_size:
                        column_samples[i].append(v)
                
                if any(v is not None for v in values):
                    rows_data.append(RowData(
                        row_index=row_idx,
                        row_data=row_dict,
                        row_hash=self._compute_row_hash(row_dict)
                    ))
            
            row_idx += 1
        
        workbook.close()
        
        column_types = {}
        for i, samples in column_samples.items():
            col_name = headers[i] if i < len(headers) else f"column_{i}"
            column_types[col_name] = self._detect_column_type(samples)
        
        info = SpreadsheetInfo(
            spreadsheet_id=str(uuid.uuid4())[:8],
            name=os.path.basename(file_path),
            source_path=file_path,
            file_type="xlsx",
            sheet_name=sheet_names[sheet_index],
            sheet_index=sheet_index,
            row_count=len(rows_data),
            column_count=len(headers),
            column_headers=headers,
            column_types=column_types,
            status="ready"
        )
        
        return info, rows_data
    
    def parse_xls(self, file_path: str, sheet_index: int = 0) -> Tuple[SpreadsheetInfo, List[RowData]]:
        if not XLRD_AVAILABLE:
            raise ImportError("xlrd required for XLS files. Install with: pip install xlrd")
        
        workbook = xlrd.open_workbook(file_path)
        
        if sheet_index >= workbook.nsheets:
            sheet_index = 0
        
        sheet = workbook.sheet_by_index(sheet_index)
        
        rows_data = []
        headers = []
        column_samples: Dict[int, List[Any]] = {}
        
        for row_idx in range(sheet.nrows):
            if row_idx >= self.max_rows:
                break
            
            values = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
            
            if row_idx == 0:
                if all(v for v in values[:min(5, len(values))]):
                    headers = [str(v).strip() if v else f"column_{i}" for i, v in enumerate(values)]
                else:
                    headers = [f"column_{i}" for i in range(len(values))]
            else:
                row_dict = {}
                for i, v in enumerate(values):
                    col_name = headers[i] if i < len(headers) else f"column_{i}"
                    if isinstance(v, float) and v.is_integer():
                        v = int(v)
                    row_dict[col_name] = v
                    if i not in column_samples:
                        column_samples[i] = []
                    if len(column_samples[i]) < self.sample_size:
                        column_samples[i].append(v)
                
                if any(v for v in values):
                    rows_data.append(RowData(
                        row_index=row_idx,
                        row_data=row_dict,
                        row_hash=self._compute_row_hash(row_dict)
                    ))
        
        column_types = {}
        for i, samples in column_samples.items():
            col_name = headers[i] if i < len(headers) else f"column_{i}"
            column_types[col_name] = self._detect_column_type(samples)
        
        info = SpreadsheetInfo(
            spreadsheet_id=str(uuid.uuid4())[:8],
            name=os.path.basename(file_path),
            source_path=file_path,
            file_type="xls",
            sheet_name=sheet.name,
            sheet_index=sheet_index,
            row_count=len(rows_data),
            column_count=len(headers),
            column_headers=headers,
            column_types=column_types,
            status="ready"
        )
        
        return info, rows_data
    
    def parse_csv(self, file_path: str, delimiter: str = ",") -> Tuple[SpreadsheetInfo, List[RowData]]:
        file_type = "tsv" if delimiter == "\t" else "csv"
        
        rows_data = []
        headers = []
        column_samples: Dict[int, List[Any]] = {}
        
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f, delimiter=delimiter)
            
            for row_idx, row in enumerate(reader):
                if row_idx >= self.max_rows + 1:
                    break
                
                if row_idx == 0:
                    headers = [v.strip() if v else f"column_{i}" for i, v in enumerate(row)]
                else:
                    row_dict = {}
                    for i, v in enumerate(row):
                        col_name = headers[i] if i < len(headers) else f"column_{i}"
                        v = v.strip()
                        if v == "":
                            v = None
                        elif v.lower() in ("true", "yes"):
                            v = True
                        elif v.lower() in ("false", "no"):
                            v = False
                        else:
                            try:
                                if "." in v:
                                    v = float(v.replace(",", ""))
                                else:
                                    v = int(v.replace(",", ""))
                            except ValueError:
                                pass
                        
                        row_dict[col_name] = v
                        if i not in column_samples:
                            column_samples[i] = []
                        if len(column_samples[i]) < self.sample_size:
                            column_samples[i].append(v)
                    
                    if any(v for v in row):
                        rows_data.append(RowData(
                            row_index=row_idx,
                            row_data=row_dict,
                            row_hash=self._compute_row_hash(row_dict)
                        ))
        
        column_types = {}
        for i, samples in column_samples.items():
            col_name = headers[i] if i < len(headers) else f"column_{i}"
            column_types[col_name] = self._detect_column_type(samples)
        
        info = SpreadsheetInfo(
            spreadsheet_id=str(uuid.uuid4())[:8],
            name=os.path.basename(file_path),
            source_path=file_path,
            file_type=file_type,
            row_count=len(rows_data),
            column_count=len(headers),
            column_headers=headers,
            column_types=column_types,
            status="ready"
        )
        
        return info, rows_data
    
    def parse_file(
        self, 
        file_path: str, 
        sheet_index: int = 0,
        delimiter: str = ","
    ) -> Tuple[SpreadsheetInfo, List[RowData]]:
        file_type = self.get_file_type(file_path)
        
        if not file_type:
            raise ValueError(f"Unsupported file type: {file_path}")
        
        if file_type == "xlsx":
            return self.parse_xlsx(file_path, sheet_index)
        elif file_type == "xls":
            return self.parse_xls(file_path, sheet_index)
        elif file_type in ("csv", "tsv"):
            actual_delimiter = "\t" if file_type == "tsv" else delimiter
            return self.parse_csv(file_path, actual_delimiter)
        else:
            raise ValueError(f"No parser for file type: {file_type}")
    
    def generate_row_text(self, row_data: RowData, headers: List[str]) -> str:
        parts = []
        for col in headers:
            val = row_data.row_data.get(col)
            if val is not None:
                parts.append(f"{col}: {val}")
        return " | ".join(parts)
    
    def to_storage_dict(self, info: SpreadsheetInfo) -> dict:
        return {
            "id": info.spreadsheet_id,
            "name": info.name,
            "source_path": info.source_path,
            "file_type": info.file_type,
            "sheet_name": info.sheet_name,
            "sheet_index": info.sheet_index,
            "row_count": info.row_count,
            "column_count": info.column_count,
            "column_headers": info.column_headers,
            "column_types": info.column_types,
            "status": info.status,
            "error_message": info.error_message,
        }
    
    def row_to_storage_dict(self, row: RowData, spreadsheet_id: str) -> dict:
        return {
            "id": str(uuid.uuid4()),
            "spreadsheet_id": spreadsheet_id,
            "row_index": row.row_index,
            "row_data": row.row_data,
            "row_hash": row.row_hash,
        }


def process_spreadsheet(
    file_path: str,
    sheet_index: int = 0,
    delimiter: str = ","
) -> Tuple[SpreadsheetInfo, List[RowData]]:
    processor = SpreadsheetProcessor()
    return processor.parse_file(file_path, sheet_index, delimiter)


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python -m src.shared.spreadsheet_processor <file_path> [sheet_index]")
        sys.exit(1)
    
    file_path = sys.argv[1]
    sheet_index = int(sys.argv[2]) if len(sys.argv) > 2 else 0
    
    try:
        info, rows = process_spreadsheet(file_path, sheet_index)
        print(f"File: {info.name}")
        print(f"Type: {info.file_type}")
        print(f"Rows: {info.row_count}")
        print(f"Columns: {info.column_count}")
        print(f"Headers: {info.column_headers}")
        print(f"Types: {info.column_types}")
        print(f"\nFirst 3 rows:")
        for row in rows[:3]:
            print(f"  Row {row.row_index}: {row.row_data}")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
